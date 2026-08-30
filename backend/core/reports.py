"""Geração de relatórios em Excel/PDF a partir de listas de dicts já
filtradas pelo controller. openpyxl e reportlab são puro-Python — sem
dependência de binário externo (wkhtmltopdf) ou libs nativas pesadas
(Cairo/Pango via weasyprint), importante na VM de produção com pouca RAM.
"""
import io
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

# Colunas de texto livre — sem isso, um valor longo sem espaços (ex.: uma
# sequência de pontos digitada sem querer) não teria onde quebrar linha e
# estouraria a largura da tabela, empurrando as colunas seguintes pra fora
# da página (bug relatado: PDF de manutenções com "Serviço Feito" longo).
_COLUNAS_LARGAS_PDF = {'descricao', 'servico_feito', 'ocorrencia', 'nome'}


def gerar_excel(titulo: str, colunas: list, linhas: list) -> io.BytesIO:
    """colunas: [(chave, rótulo), ...]. linhas: lista de dicts."""
    wb = Workbook()
    ws = wb.active
    ws.title = titulo[:31]  # limite do Excel pro nome da aba

    for col_idx, (_, rotulo) in enumerate(colunas, start=1):
        cell = ws.cell(row=1, column=col_idx, value=rotulo)
        cell.font = Font(bold=True)

    for row_idx, linha in enumerate(linhas, start=2):
        for col_idx, (chave, _) in enumerate(colunas, start=1):
            ws.cell(row=row_idx, column=col_idx, value=linha.get(chave, ''))

    for col_idx, (chave, rotulo) in enumerate(colunas, start=1):
        largura = max(len(rotulo), *(len(str(linha.get(chave, ''))) for linha in linhas)) if linhas else len(rotulo)
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = min(largura + 2, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def gerar_pdf(titulo: str, colunas: list, linhas: list) -> io.BytesIO:
    """colunas: [(chave, rótulo), ...]. linhas: lista de dicts."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4),
                            topMargin=1.5 * cm, bottomMargin=1.5 * cm,
                            leftMargin=1.5 * cm, rightMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    cell_style = ParagraphStyle('cell', parent=styles['Normal'], fontSize=8, leading=10)

    # Colunas de texto livre ganham mais espaço; o resto (número, status,
    # valor, datas...) fica estreito — evita que uma coluna de texto longo
    # domine a largura da página inteira.
    pesos = [3 if chave in _COLUNAS_LARGAS_PDF else 1 for chave, _ in colunas]
    largura_disponivel = landscape(A4)[0] - 3 * cm  # descontando as margens
    total_peso = sum(pesos)
    col_widths = [largura_disponivel * peso / total_peso for peso in pesos]

    dados = [[rotulo for _, rotulo in colunas]]
    for linha in linhas:
        dados.append([Paragraph(escape(str(linha.get(chave, '')) or ''), cell_style) for chave, _ in colunas])

    tabela = Table(dados, colWidths=col_widths, repeatRows=1)
    tabela.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
    ]))

    doc.build([Paragraph(titulo, styles['Title']), tabela])
    buffer.seek(0)
    return buffer
